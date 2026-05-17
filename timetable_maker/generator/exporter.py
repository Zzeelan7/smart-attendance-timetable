"""
generator/exporter.py
Exports the generated timetable to Excel (.xlsx) and PDF.
"""

import os
import io
from datetime import datetime
from .data import DAYS, PERIOD_TIMES, SUBJECT_COLORS

# ── Excel Export ──────────────────────────────────────────────────

def export_excel(result: dict, semester: int = 4) -> bytes:
    """Returns the Excel file as bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import (PatternFill, Font, Alignment,
                                  Border, Side)
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    period_labels = ['P1\n9:00–10:00', 'P2\n10:00–11:00',
                     'P3\n11:15–12:15', 'P4\n12:15–1:15',
                     'P5\n2:00–3:00', 'P6\n3:00–4:00',
                     'P7\n4:00–5:00 (DIP MATHS/NCMC)']

    def safe_color(raw, default='334155'):
        """Return a clean 6-char RRGGBB hex string for openpyxl."""
        if not raw:
            return default
        return str(raw).lstrip('#')[:6].ljust(6, '0')

    def make_sheet(ws, section, grid, title):
        ws.title = title
        thin = Side(style='thin', color='222222')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Title row
        ws.merge_cells('A1:I1')
        ws['A1'] = f'4th Semester ECE — Section {section} Timetable'
        ws['A1'].font = Font(bold=True, size=13, color='FFFFFF')
        ws['A1'].fill = PatternFill('solid', fgColor='1a2235')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28

        # Sub-title
        ws.merge_cells('A2:I2')
        ws['A2'] = f'Generated: {datetime.now().strftime("%d-%m-%Y %H:%M")}   |   Academic Year 2025-26'
        ws['A2'].font = Font(italic=True, size=9, color='AAAAAA')
        ws['A2'].fill = PatternFill('solid', fgColor='0f1525')
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[2].height = 18

        # Header row
        headers = ['Day', 'P1 (9-10)', 'P2 (10-11)', 'P3 (11:15-12:15)',
                   'P4 (12:15-1:15)', 'P5 (2-3)', 'P6 (3-4)', 'P7 (4-5) DIP/NCMC', 'Remarks']
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = Font(bold=True, color='FFFFFF', size=9)
            cell.fill = PatternFill('solid', fgColor='161d30')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
            ws.row_dimensions[3].height = 30

        # Column widths
        ws.column_dimensions['A'].width = 12
        for col in range(2, 10):
            ws.column_dimensions[get_column_letter(col)].width = 18

        # Data rows
        for day_idx, day in enumerate(DAYS):
            row = day_idx + 4
            ws.row_dimensions[row].height = 45
            # Day cell
            day_cell = ws.cell(row=row, column=1, value=day)
            day_cell.font = Font(bold=True, color='FFFFFF', size=9)
            day_cell.fill = PatternFill('solid', fgColor='161d30')
            day_cell.alignment = Alignment(horizontal='center', vertical='center')
            day_cell.border = border

            for p in range(6):
                cell_data = grid[day_idx][p]
                col = p + 2
                xls_cell = ws.cell(row=row, column=col)
                xls_cell.border = border
                xls_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                if cell_data:
                    display   = cell_data.get('display') or ''
                    teacher   = cell_data.get('teacher') or ''
                    room      = cell_data.get('room') or ''
                    color_hex = safe_color(cell_data.get('color'), '334155')
                    text = display
                    if teacher and cell_data.get('type') != 'fixed':
                        text += f'\n{teacher}'
                    if room:
                        text += f'\nRoom {room}'
                    xls_cell.value = text
                    xls_cell.font = Font(bold=True, size=8, color='FFFFFF')
                    xls_cell.fill = PatternFill('solid', fgColor=color_hex)
                else:
                    xls_cell.value = '—'
                    xls_cell.font = Font(size=8, color='666666')
                    xls_cell.fill = PatternFill('solid', fgColor='0a0e1a')

            # P7 cell (always blocked)
            p7_cell = ws.cell(row=row, column=8, value='DIP MATHS\n(Diploma Students)')
            p7_cell.font = Font(size=8, italic=True, color='888888')
            p7_cell.fill = PatternFill('solid', fgColor='111827')
            p7_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            p7_cell.border = border

        # Saturday row
        sat_row = 9
        ws.row_dimensions[sat_row].height = 40
        sat_cell = ws.cell(row=sat_row, column=1, value='Saturday')
        sat_cell.font = Font(bold=True, color='FFFFFF', size=9)
        sat_cell.fill = PatternFill('solid', fgColor='161d30')
        sat_cell.alignment = Alignment(horizontal='center', vertical='center')
        sat_cell.border = border
        sat_fixed = [
            (2, 3, 'MINI PROJECT', '475569'),
            (4, 5, 'PROCTORING', '374151'),
            (6, 7, 'EXCL (Elective)', '1e293b'),
            (8, 8, 'DIP MATHS', '111827'),
        ]
        for c1, c2, txt, clr in sat_fixed:
            if c1 == c2:
                xc = ws.cell(row=sat_row, column=c1, value=txt)
                xc.fill = PatternFill('solid', fgColor=clr)
                xc.font = Font(bold=True, size=8, color='AAAAAA')
                xc.alignment = Alignment(horizontal='center', vertical='center')
                xc.border = border
            else:
                ws.merge_cells(start_row=sat_row, start_column=c1,
                                end_row=sat_row, end_column=c2)
                xc = ws.cell(row=sat_row, column=c1, value=txt)
                xc.fill = PatternFill('solid', fgColor=clr)
                xc.font = Font(bold=True, size=9, color='AAAAAA')
                xc.alignment = Alignment(horizontal='center', vertical='center')
                xc.border = border

    # Create section sheets
    for sec in ['A', 'B']:
        ws = wb.create_sheet(title=f'Section {sec}')
        make_sheet(ws, sec, result[sec], f'Section {sec}')

    # Teacher timetable sheets
    for teacher, t_grid in result.get('teacher_schedules', {}).items():
        ws = wb.create_sheet(title=teacher[:28])
        ws['A1'] = f'Teacher Timetable: {teacher}'
        ws['A1'].font = Font(bold=True, size=11, color='FFFFFF')
        ws['A1'].fill = PatternFill('solid', fgColor='1a2235')
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 24

        thin = Side(style='thin', color='222222')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = ['Day', 'P1', 'P2', 'P3', 'P4', 'P5', 'P6']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=2, column=col, value=h)
            c.font = Font(bold=True, color='FFFFFF', size=9)
            c.fill = PatternFill('solid', fgColor='161d30')
            c.alignment = Alignment(horizontal='center')
            c.border = border
            ws.column_dimensions[get_column_letter(col)].width = 16

        for day_idx, day in enumerate(DAYS):
            row = day_idx + 3
            ws.row_dimensions[row].height = 38
            dc = ws.cell(row=row, column=1, value=day)
            dc.font = Font(bold=True, color='FFFFFF', size=9)
            dc.fill = PatternFill('solid', fgColor='161d30')
            dc.alignment = Alignment(horizontal='center', vertical='center')
            dc.border = border

            for p in range(6):
                cell_data = t_grid[day_idx][p]
                col = p + 2
                xc = ws.cell(row=row, column=col)
                xc.border = border
                xc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                if cell_data:
                    txt = f"{cell_data.get('display','')}"
                    sec_lbl = cell_data.get('section', '')
                    if sec_lbl:
                        txt += f"\nSec {sec_lbl}"
                    if cell_data.get('room'):
                        txt += f"\nRoom {cell_data['room']}"
                    xc.value = txt
                    xc.font = Font(bold=True, size=8, color='FFFFFF')
                    xc.fill = PatternFill('solid',
                        fgColor=safe_color(cell_data.get('color'), '334155'))
                else:
                    xc.value = '—'
                    xc.font = Font(size=8, color='555555')
                    xc.fill = PatternFill('solid', fgColor='0a0e1a')

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF Export ────────────────────────────────────────────────────

def export_pdf(result: dict, semester: int = 4) -> bytes:
    """Returns the PDF as bytes using ReportLab."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story = []

    period_headers = ['Day', 'P1\n9:00', 'P2\n10:00',
                      'P3\n11:15', 'P4\n12:15',
                      'P5\n2:00', 'P6\n3:00', 'P7\n4:00']

    def hex_to_rl(h):
        h = h.lstrip('#')
        r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
        return colors.Color(r/255, g/255, b/255)

    title_style = ParagraphStyle('title', fontSize=14, fontName='Helvetica-Bold',
                                  textColor=colors.HexColor('#4f8ef7'), spaceAfter=4)
    sub_style = ParagraphStyle('sub', fontSize=8, fontName='Helvetica',
                                textColor=colors.HexColor('#8fa3c8'), spaceAfter=12)
    cell_style = ParagraphStyle('cell', fontSize=7, fontName='Helvetica-Bold',
                                 textColor=colors.white, alignment=1)

    for sec in ['A', 'B']:
        story.append(Paragraph(
            f'4th Semester ECE — Section {sec} Timetable', title_style))
        story.append(Paragraph(
            f'Generated: {datetime.now().strftime("%d-%m-%Y %H:%M")}  |  Academic Year 2025-26',
            sub_style))

        data = [period_headers]
        for day_idx, day in enumerate(DAYS):
            row = [Paragraph(f'<b>{day}</b>',
                             ParagraphStyle('d', fontSize=8, fontName='Helvetica-Bold',
                                             textColor=colors.white, alignment=1))]
            for p in range(6):
                cell_data = result[sec][day_idx][p]
                if cell_data:
                    txt = cell_data.get('display', '')
                    t = cell_data.get('teacher', '')
                    if t and cell_data.get('type') not in ('fixed',):
                        txt += f'<br/><font size="6">{t}</font>'
                    row.append(Paragraph(txt, cell_style))
                else:
                    row.append(Paragraph('—', ParagraphStyle('e', fontSize=7,
                                          textColor=colors.HexColor('#555555'), alignment=1)))
            # P7 always DIP MATHS
            row.append(Paragraph('<font size="6">DIP MATHS</font>',
                                  ParagraphStyle('p7', fontSize=6, textColor=colors.HexColor('#555555'), alignment=1)))
            data.append(row)

        # Saturday
        sat_row = [Paragraph('<b>Saturday</b>', ParagraphStyle('d', fontSize=8,
                              fontName='Helvetica-Bold', textColor=colors.white, alignment=1)),
                   Paragraph('MINI PROJECT', cell_style),
                   Paragraph('MINI PROJECT', cell_style),
                   Paragraph('PROCTORING', cell_style),
                   Paragraph('PROCTORING', cell_style),
                   Paragraph('EXCL', cell_style),
                   Paragraph('EXCL', cell_style),
                   Paragraph('—', cell_style)]
        data.append(sat_row)

        col_widths = [2.5*cm] + [3.3*cm]*7
        table = Table(data, colWidths=col_widths, repeatRows=1)

        ts = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#161d30')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#2d3748')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1),
             [colors.HexColor('#0f1525'), colors.HexColor('#0a0e1a')]),
            ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#161d30')),
        ]
        # Color each cell based on subject
        for day_idx in range(len(DAYS)):
            for p in range(6):
                cell_data = result[sec][day_idx][p]
                if cell_data and cell_data.get('color'):
                    c = hex_to_rl(cell_data['color'])
                    ts.append(('BACKGROUND', (p+1, day_idx+1), (p+1, day_idx+1), c))

        table.setStyle(TableStyle(ts))
        story.append(table)
        story.append(Spacer(1, 1*cm))

    doc.build(story)
    return buf.getvalue()
